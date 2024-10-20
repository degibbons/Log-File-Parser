import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fdlg
import re
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import shutil



class MainWindow(tk.Tk):
    """class representing the overall gui window"""

    def __init__(self):
        super().__init__()
        self.title("Log File Parser")
        # self.geometry("1250x780+20+20")
        self.selected_log_source_file = ""
        self.target_archive_file = ""
        self.target_dir = ""
        self.log_file_text = ""
        self.data_dict = {}

        # pinkStyle = ttk.Style()
        # pinkStyle.configure("test0Style.TFrame", background="pink")
        # greenStyle = ttk.Style()
        # greenStyle.configure("test1Style.TFrame", background="purple")
        # blueStyle = ttk.Style()
        # blueStyle.configure("test2Style.TFrame", background="orange")
        # purpleStyle = ttk.Style()
        # purpleStyle.configure("test3Style.TFrame", background="brown")

        # self.rowconfigure(0, weight=1)
        # self.columnconfigure(0, weight=1)

        # Title Frame
        title_frame = ttk.Frame(
            master=self,
            # style='test1Style.TFrame'
        )
        title_frame.grid(column=0, row=0, columnspan=1, rowspan=1, sticky=tk.W + tk.E)
        title_frame.columnconfigure(0, weight=1)
        # title_frame.columnconfigure(4,weight=1)
        # title_frame.rowconfigure(0,weight=1)
        self.main_title_label = ttk.Label(
            master=title_frame,
            text="Log File Parse Tool",
            relief=tk.FLAT,
            font=("Times New Roman", 18, "bold"),
            #  background="yellow"
        )
        self.main_title_label.grid(
            row=0, column=0, columnspan=4, rowspan=1, ipadx=5, pady=0, sticky=tk.S
        )
        self.version_title_label = ttk.Label(
            master=title_frame,
            text="v2.0.1",
            font=("Times New Roman", 8, "italic"),
            # background="red"
        )
        self.version_title_label.grid(
            row=1, column=0, columnspan=4, rowspan=1, padx=0, pady=0, sticky=tk.N
        )
        self.help_button = ttk.Button(
            master=title_frame, text="Help", command=self.open_help_box
        )
        self.help_button.grid(
            row=1, column=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.E
        )

        sep_1 = ttk.Separator(master=self, orient="horizontal")
        sep_1.grid(row=1, column=0, sticky="nesw")

        # Start Frame
        start_frame = ttk.Frame(
            master=self,
            # style='test0Style.TFrame'
        )
        start_frame.grid(
            column=0, row=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.E
        )
        start_frame.columnconfigure(1, weight=1)
        self.select_destination_file_button = ttk.Button(
            master=start_frame,
            text="Select Destination File",
            command=self.select_destination_file,
        )
        self.select_destination_file_button.grid(
            row=0, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )

        self.select_destination_file_label_frame = ttk.LabelFrame(
            master=start_frame,
            # labelwidget=self.select_destination_file_label,
            text="Destination File Name",
            height=2,
        )
        self.select_destination_file_label = ttk.Label(
            master=self.select_destination_file_label_frame,
            text="",
            # background="red",
            width=130,
            anchor=tk.W,
        )
        self.select_destination_file_label_frame.grid(
            row=0, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.E
        )
        self.select_destination_file_label.grid(
            row=0,
            column=0,
            columnspan=1,
            rowspan=1,
            padx=5,
            pady=5,
            sticky=tk.N + tk.W + tk.E,
        )
        self.indicator_light_destination_file = tk.Canvas(
            master=start_frame, bg="white", height=50, width=100
        )
        self.indicator_light_destination_file.grid(
            column=5, row=0, columnspan=1, rowspan=1, sticky=tk.E
        )
        self.indicator_light_destination_file.create_text(
            30, 25, text="Selection\nIndicator"
        )
        self.destination_file_light = self.indicator_light_destination_file.create_oval(
            75, 20, 90, 35, fill="red"
        )

        self.destinationfilekeep_yesno = tk.IntVar()
        self.keep_destination_file_checkbutton = ttk.Checkbutton(
            master=start_frame,
            text="Keep when reset?",
            onvalue=1,
            offvalue=0,
            variable=self.destinationfilekeep_yesno,
        )
        self.keep_destination_file_checkbutton.grid(
            row=0,
            column=6,
            columnspan=1,
            rowspan=1,
            padx=5,
            pady=5,
            sticky=tk.W + tk.N + tk.S,
        )
        self.destinationfilekeep_yesno.set(value=0)

        self.select_log_file_button = ttk.Button(
            master=start_frame,
            text="Select Log Source File",
            command=self.select_log_source_file,
        )
        self.select_log_file_button.grid(
            row=1, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.select_log_file_label_frame = ttk.LabelFrame(
            master=start_frame,
            # labelwidget=self.select_destination_file_label,
            text="Log File Name",
            height=2,
        )
        self.select_log_file_label = ttk.Label(
            master=self.select_log_file_label_frame,
            text="",
            # background="green",
            width=130,
            # height=2
        )
        self.select_log_file_label_frame.grid(
            row=1, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.E
        )
        self.select_log_file_label.grid(
            row=0,
            column=0,
            columnspan=1,
            rowspan=1,
            padx=5,
            pady=5,
            sticky=tk.N + tk.W + tk.E,
        )
        self.import_data_button = ttk.Button(
            master=start_frame, text="Parse Data", command=self.parse_data
        )
        self.import_data_button.grid(
            row=1,
            column=6,
            columnspan=1,
            rowspan=1,
            padx=5,
            pady=5,
            # sticky=tk.E + tk.W,
        )
        self.indicator_light_importdata = tk.Canvas(
            master=start_frame, bg="white", height=50, width=100
        )
        self.indicator_light_importdata.grid(
            column=5, row=1, columnspan=1, rowspan=1, sticky=tk.E
        )
        self.indicator_light_importdata.create_text(30, 25, text="Selection\nIndicator")
        self.import_data_light = self.indicator_light_importdata.create_oval(
            75, 20, 90, 35, fill="red"
        )
        self.shift_log_file_button = ttk.Button(
            master=start_frame,
            text="Select Log File Destination",
            command=self.shift_log_file,
        )
        self.shift_log_file_button.grid(
            row=2, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.E
        )
        self.shift_log_file_label_frame = ttk.LabelFrame(
            master=start_frame,
            # labelwidget=self.select_destination_file_label,
            text="Log File Destination",
            height=2,
        )
        self.shift_log_file_label = ttk.Label(
            master=self.shift_log_file_label_frame,
            text="",
            # background="blue",
            width=130,
            # height=2
        )
        self.shift_log_file_label_frame.grid(
            row=2, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.E
        )
        self.shift_log_file_label.grid(
            row=0, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.E
        )
        self.indicator_light_shift_logfile = tk.Canvas(
            master=start_frame, bg="white", height=50, width=100
        )
        self.indicator_light_shift_logfile.grid(
            row=2, column=5, columnspan=1, rowspan=1, sticky=tk.E
        )
        self.indicator_light_shift_logfile.create_text(
            30, 25, text="Selection\nIndicator"
        )
        self.shift_log_file_light = self.indicator_light_shift_logfile.create_oval(
            75, 20, 90, 35, fill="red"
        )

        self.destinationdirectorykeep_yesno = tk.IntVar()
        self.keep_destination_directory_checkbutton = ttk.Checkbutton(
            master=start_frame,
            text="Keep when reset?",
            onvalue=1,
            offvalue=0,
            variable=self.destinationdirectorykeep_yesno,
        )
        self.keep_destination_directory_checkbutton.grid(
            row=2,
            column=6,
            columnspan=1,
            rowspan=1,
            padx=5,
            pady=5,
            sticky=tk.W + tk.N + tk.S,
        )
        self.destinationdirectorykeep_yesno.set(value=0)

        sep_2 = ttk.Separator(master=self, orient="horizontal")
        sep_2.grid(row=2, column=0, sticky="nesw")

        # Input Frame
        input_frame = ttk.Frame(
            master=self,
            # style='test2Style.TFrame'
        )
        # label_font=font(family="Times New Roman", size=9, weight="bold")
        input_frame.grid(
            column=0,
            row=2,
            columnspan=1,
            rowspan=1,
            padx=5,
            pady=5,
            sticky=tk.W + tk.E + tk.N + tk.S,
        )

        input_frame.columnconfigure(0, weight=1)
        input_frame.columnconfigure(1, weight=1)
        input_frame.columnconfigure(2, weight=1)
        input_frame.columnconfigure(3, weight=1)
        input_frame.columnconfigure(4, weight=1)

        self.proposal_number_label = ttk.Label(
            master=input_frame,
            text="Proposal #:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.proposal_number_label.grid(
            row=0, column=0, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W + tk.S
        )
        self.proposalnumber_textvar = tk.StringVar()
        self.proposal_number_input = ttk.Entry(
            master=input_frame, textvariable=self.proposalnumber_textvar, width=35
        )
        self.proposal_number_input.grid(
            row=1, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.N
        )
        self.proposalnumber_textvar.set(value="Enter Here")
        self.name_label = ttk.Label(
            master=input_frame,
            text="Name:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.name_label.grid(
            row=0, column=1, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W + tk.S
        )
        self.name_textvar = tk.StringVar()
        self.name_input = ttk.Entry(
            master=input_frame, textvariable=self.name_textvar, width=35
        )
        self.name_input.grid(
            row=1, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.N
        )
        self.name_textvar.set(value="Enter Here")
        self.date_label = ttk.Label(
            master=input_frame,
            text="Date:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.date_label.grid(
            row=0, column=2, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W + tk.S
        )
        self.date_textvar = tk.StringVar()
        self.date_input = ttk.Entry(
            master=input_frame, textvariable=self.date_textvar, width=35
        )
        self.date_input.grid(
            row=1, column=2, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.N
        )
        self.date_textvar.set(value="Known")

        self.position_label = ttk.Label(
            master=input_frame,
            text="Position:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.position_label.grid(
            row=0, column=3, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.position_textvar = tk.StringVar()
        self.position_input = ttk.Combobox(
            master=input_frame,
            textvariable=self.position_textvar,
            values=(
                "Faculty",
                "Staff",
                "Student",
                "Other",
                "Faculty/Staff",
                "Faculty/Student",
                "Faculty/Other",
                "Staff/Student",
                "Staff/Other",
                "Student/Other",
                "Faculty/Staff/Student",
                "Faculty/Staff/Other",
                "Faculty/Student/Other",
                "Staff/Student/Other",
                "Faculty/Staff/Student/Other",
            ),
        )
        # self.position_input = ttk.Entry(
        #     master=input_frame, textvariable=self.position_textvar, width=35
        # )
        self.position_input.grid(
            row=1, column=3, columnspan=1, rowspan=1, padx=2, pady=5, sticky=tk.W + tk.E
        )
        self.position_textvar.set(value="Select Here")

        self.operator_label = ttk.Label(
            master=input_frame,
            text="Operator:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.operator_label.grid(
            row=0, column=4, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.operator_textvar = tk.StringVar()
        self.operator_input = ttk.Entry(
            master=input_frame, textvariable=self.operator_textvar, width=35
        )
        self.operator_input.grid(
            row=1, column=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.operator_textvar.set(value="Select Here")
        self.nyit_or_ext_label = ttk.Label(
            master=input_frame,
            text="Internal/External/Both:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.nyit_or_ext_label.grid(
            row=2, column=0, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.nyitorext_textvar = tk.StringVar()
        self.nyit_or_ext_input = ttk.Combobox(
            master=input_frame,
            textvariable=self.nyitorext_textvar,
            width=35,
            values=(
                "Internal",
                "External",
                "Both",
            ),
        )
        self.nyit_or_ext_input.grid(
            row=3, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.nyitorext_textvar.set(value="Enter Here")
        self.filename_label = ttk.Label(
            master=input_frame,
            text="Filename:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.filename_label.grid(
            row=2, column=1, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W + tk.E
        )
        self.filename_textvar = tk.StringVar()
        self.filename_input = ttk.Entry(
            master=input_frame,
            textvariable=self.filename_textvar,
            width=35,
            # background="red"
        )
        self.filename_input.grid(
            row=3, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.filename_textvar.set(value="Known")
        self.voltage_label = ttk.Label(
            master=input_frame,
            text="Voltage:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.voltage_label.grid(
            row=2, column=2, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.voltage_textvar = tk.StringVar()
        self.voltage_input = ttk.Entry(
            master=input_frame, textvariable=self.voltage_textvar, width=20
        )
        self.voltage_input.grid(
            row=3, column=2, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.voltage_textvar.set(value="Known")
        self.current_label = ttk.Label(
            master=input_frame,
            text="Current:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.current_label.grid(
            row=2, column=3, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.current_textvar = tk.StringVar()
        self.current_input = ttk.Entry(
            master=input_frame, textvariable=self.current_textvar, width=20
        )
        self.current_input.grid(
            row=3, column=3, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.current_textvar.set(value="Known")
        self.resolution_label = ttk.Label(
            master=input_frame,
            text="Resolution:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.resolution_label.grid(
            row=2, column=4, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.resolution_textvar = tk.StringVar()
        self.resolution_combobox = ttk.Combobox(
            master=input_frame,
            textvariable=self.resolution_textvar,
            values=("0.5K", "1K", "2K"),
        )
        self.resolution_combobox.grid(
            row=3, column=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.N
        )
        self.resolution_textvar.set(value="Select Here")
        self.filter_label = ttk.Label(
            master=input_frame,
            text="Filter:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.filter_label.grid(
            row=4, column=0, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.filter_textvar = tk.StringVar()
        self.filter_combobox = ttk.Combobox(
            master=input_frame,
            textvariable=self.filter_textvar,
            values=("Br 0.25 mm", "AL 1.0 mm", "None"),
        )
        self.filter_combobox.grid(
            row=5, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.filter_textvar.set(value="Enter or Select Here")
        self.exposure_label = ttk.Label(
            master=input_frame,
            text="Exposure:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.exposure_label.grid(
            row=4, column=1, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.exposure_textvar = tk.StringVar()
        self.exposure_input = ttk.Entry(
            master=input_frame, textvariable=self.exposure_textvar, width=20
        )
        self.exposure_input.grid(
            row=5, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.exposure_textvar.set(value="Known")
        self.pixel_size_label = ttk.Label(
            master=input_frame,
            text="Pixel Size:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.pixel_size_label.grid(
            row=4, column=2, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.pixelsize_textvar = tk.StringVar()
        self.pixel_size_input = ttk.Entry(
            master=input_frame, textvariable=self.pixelsize_textvar, width=20
        )
        self.pixel_size_input.grid(
            row=5, column=2, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.pixelsize_textvar.set(value="Known")
        self.image_format_label = ttk.Label(
            master=input_frame,
            text="Image Format:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.image_format_label.grid(
            row=4, column=3, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.imageformat_textvar = tk.StringVar()
        self.image_format_input = ttk.Entry(
            master=input_frame, textvariable=self.imageformat_textvar, width=20
        )
        self.image_format_input.grid(
            row=5, column=3, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.imageformat_textvar.set(value="Known")
        self.type_of_scan_label = ttk.Label(
            master=input_frame,
            text="Type of Scan:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.type_of_scan_label.grid(
            row=4, column=4, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.typeofscan_textvar = tk.StringVar()
        self.type_of_scan_input = ttk.Entry(
            master=input_frame, textvariable=self.typeofscan_textvar, width=35
        )
        self.type_of_scan_input.grid(
            row=5, column=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.typeofscan_textvar.set(value="Enter Here")
        self.rotation_step_label = ttk.Label(
            master=input_frame,
            text="Rotation Step:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.rotation_step_label.grid(
            row=6, column=0, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.rotationstep_textvar = tk.StringVar()
        self.rotation_step_input = ttk.Entry(
            master=input_frame, textvariable=self.rotationstep_textvar, width=35
        )
        self.rotation_step_input.grid(
            row=7, column=0, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.rotationstep_textvar.set(value="Known")
        self.frames_label = ttk.Label(
            master=input_frame,
            text="Frames:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.frames_label.grid(
            row=6, column=1, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.frames_textvar = tk.StringVar()
        self.frames_input = ttk.Entry(
            master=input_frame, textvariable=self.frames_textvar, width=20
        )
        self.frames_input.grid(
            row=7, column=1, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.frames_textvar.set(value="Known")
        self.random_movement_label = ttk.Label(
            master=input_frame,
            text="Random Movement:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.random_movement_label.grid(
            row=6, column=2, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.randommovement_textvar = tk.StringVar()
        self.random_movement_input = ttk.Entry(
            master=input_frame, textvariable=self.randommovement_textvar
        )
        self.random_movement_input.grid(
            row=7, column=2, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.randommovement_textvar.set(value="Known")
        self.rotate360_yesno = tk.IntVar()
        self.rotate_360_checkbutton = ttk.Checkbutton(
            master=input_frame,
            text="Rotate 360?",
            onvalue=1,
            offvalue=0,
            variable=self.rotate360_yesno,
        )
        self.rotate360_yesno.set(value=0)
        self.rotate_360_checkbutton.grid(
            row=7, column=3, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.scan_duration_label = ttk.Label(
            master=input_frame,
            text="Scan Duration:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.scan_duration_label.grid(
            row=6, column=4, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        self.scanduration_textvar = tk.StringVar()
        self.scan_duration_input = ttk.Entry(
            master=input_frame, textvariable=self.scanduration_textvar, width=25
        )
        self.scan_duration_input.grid(
            row=7, column=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.scanduration_textvar.set(value="Known")
        self.comments_label = ttk.Label(
            master=input_frame,
            text="Comments:",
            font=("Times New Roman", 9, "bold underline"),
        )
        self.comments_label.grid(
            row=8, column=0, columnspan=1, rowspan=1, padx=5, pady=0, sticky=tk.W
        )
        # self.comments_textvar = tk.StringVar()
        self.comments_input = tk.Text(
            master=input_frame,
            height=4,
            width=72,
        )
        self.comments_input.grid(
            row=9, column=0, columnspan=3, rowspan=1, padx=5, pady=5, sticky=tk.W
        )
        self.comments_input.insert(tk.END, "Enter Here")
        # self.comments_textvar.set(value="Enter Here")
        self.logfilearchive_yesno = tk.IntVar()
        self.log_file_archived_checkbutton = ttk.Checkbutton(
            master=input_frame,
            text="Log File Archived?",
            onvalue=1,
            offvalue=0,
            variable=self.logfilearchive_yesno,
        )
        self.logfilearchive_yesno.set(value=0)
        self.log_file_archived_checkbutton.grid(
            row=9, column=3, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.N
        )
        self.visstorage_yesno = tk.IntVar()
        self.added_to_viscenter_storage_checkbutton = ttk.Checkbutton(
            master=input_frame,
            text="Added to Viscenter Storage?",
            onvalue=1,
            offvalue=0,
            variable=self.visstorage_yesno,
        )
        self.added_to_viscenter_storage_checkbutton.grid(
            row=9, column=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.W + tk.N
        )
        self.visstorage_yesno.set(value=0)

        sep_3 = ttk.Separator(master=self, orient="horizontal")
        sep_3.grid(row=3, column=0, sticky="nesw")

        # Output Frame
        output_frame = ttk.Frame(
            master=self,
            # style='test3Style.TFrame'
        )
        output_frame.grid(
            column=0, row=4, columnspan=1, rowspan=1, padx=5, pady=5, sticky=tk.E
        )
        self.selectedfile_radbuttonvar = tk.IntVar()
        self.excel_select_radiobutton = ttk.Radiobutton(
            master=output_frame,
            text="Excel File",
            variable=self.selectedfile_radbuttonvar,
            value=1,
        )
        self.gsheet_select_radiobutton = ttk.Radiobutton(
            master=output_frame,
            text="Google Sheet",
            variable=self.selectedfile_radbuttonvar,
            value=0,
        )
        self.selectedfile_radbuttonvar.set(1)
        self.excel_select_radiobutton.grid(
            row=0,
            column=0,
            columnspan=1,
            rowspan=2,
            padx=5,
            pady=5,
            sticky=tk.W,
        )
        self.gsheet_select_radiobutton.grid(
            row=2,
            column=0,
            columnspan=1,
            rowspan=2,
            padx=5,
            pady=5,
            sticky=tk.W,
        )
        self.publish_info_button = ttk.Button(
            master=output_frame, text="Publish Info", command=self.publish_info
        )
        self.publish_info_button.grid(
            row=0, column=1, columnspan=1, rowspan=2, padx=5, pady=5, sticky=tk.E
        )
        self.reset_clear_button = ttk.Button(
            master=output_frame, text="Reset / Clear", command=self.reset_clear
        )
        self.reset_clear_button.grid(
            row=2, column=1, columnspan=1, rowspan=2, padx=5, pady=5, sticky=tk.E
        )

    def open_help_box(self):
        """opens/creates a help box for reading instructions on how to use the program"""
        self.hb_object_box = tk.Toplevel(
            self
        )  # Create a new window for the help box object
        self.hb_object_box.geometry(
            "710x480"
        )  # Set the geometry of the help box window
        self.hb_object_box.title(
            "Help Box Window"
        )  # Set the title of the help box window
        frame_12 = ttk.Frame(
            master=self.hb_object_box
        )  # Create a frame to hold the page change buttons
        help_notebook = ttk.Notebook(master=self.hb_object_box)
        help_frame_1 = ttk.Frame(master=help_notebook)
        help_frame_2 = ttk.Frame(master=help_notebook)
        help_frame_3 = ttk.Frame(master=help_notebook)
        help_frame_4 = ttk.Frame(master=help_notebook)
        help_frame_5 = ttk.Frame(master=help_notebook)
        help_frame_6 = ttk.Frame(master=help_notebook)

        help_notebook.grid(row=0, column=0, padx=5, pady=5)
        help_frame_1.grid(row=0, column=0, padx=5, pady=5)
        help_frame_2.grid(row=0, column=0, padx=5, pady=5)
        help_frame_3.grid(row=0, column=0, padx=5, pady=5)
        help_frame_4.grid(row=0, column=0, padx=5, pady=5)
        help_frame_5.grid(row=0, column=0, padx=5, pady=5)
        help_frame_6.grid(row=0, column=0, padx=5, pady=5)

        help_notebook.add(help_frame_1, text="General Info")
        help_notebook.add(help_frame_2, text="Destination File")
        help_notebook.add(help_frame_3, text="Log File")
        help_notebook.add(help_frame_4, text="Parsed Information")
        help_notebook.add(help_frame_5, text="Publishing Data")
        help_notebook.add(help_frame_6, text="Other")

        general_info_label = ttk.Label(
            master=help_frame_1,
            text="""

            This is the Log File Parser Tool
            Designed and Created by Dan Gibbons
            Property of NYIT
                                       
            The purpose of this app is to make it easy to transfer log data from log files generated after use of the 
            MicroCT machine to the designated archive file to make further analysis and visualization possible.
            
            For questions about this application not addressed in this help section or 
            too complicated or underexplained, please contact dgibbo03@nyit.edu.

            For questions about data usage and publication queries, please contact khurdle@nyit.edu.
            """,
        )
        general_info_label.grid(row=0, column=0, padx=5, pady=5)

        destination_file_label = ttk.Label(
            master=help_frame_2,
            text="""

            The first step of the process is selecting the Archive Target File (.xlsx)
            file where which the data extracted will end up being published. This can be selected by 
            clicking the 'Select Destination File' button at the top left of the GUI. After clicking, 
            another native GUI should pop up prompting you to navigate and select the file.
            
            After selecting the archive file, the text display next to the button should display the 
            global pathway to the now selected document. The indicator light to the right will now turn 
            green, indicating selection has been carried out.
            
            NOTE: There is a radio button to the right that says 'Keep When Reset?'.
            This may be selected when processing numerous log files when the target archive file is 
            to remain the same for all of the log files. This simply stops the need to continue selecting 
            the archive file for every session.""",
        )
        destination_file_label.grid(row=0, column=0, padx=5, pady=5)

        log_file_label = ttk.Label(
            master=help_frame_3,
            text="""
            The second step of the log file parser process is selecting the log file from which you will be 
            extracting the usage information from each scan.
            
            Just like the previous process, the second step will require you to click the button labeled 'Select Log Source File'.
            A similar box to step 1 will pop up, prompting you for a log file to be selected. 
            Also once this is selected, it's global pathway will be displayed to the right of the button as well as the indicator
            light turning green.
            
            Only after the file is selected and the indicator light is turned green can you click the 'Parse Data' button.
            
            After parsing, the appropriate fields on the GUI (namely the ones with the word 'known' in them) will populate 
            with data from the log file selected.""",
        )
        log_file_label.grid(row=0, column=0, padx=5, pady=5)

        parsed_info_label = ttk.Label(
            master=help_frame_4,
            text="""

            The third step of the process is filling in the missing info using the log source file.

            After the proper data has been parsed from the source log file, entry fields with the 'Enter Here' text or checkboxes 
            will need to be filled in manually using either the data that could not be directly parsed from the log file itself, 
            or user known data that may not be recorded in the log source file. 

            Failure to fill an input field will result in the recorded data displaying 'Entere Here', False, or nothing.

            For specifics on field input, refer to the 'Other' tab above.
            """,
        )
        parsed_info_label.grid(row=0, column=0, padx=5, pady=5)

        publishing_data_label = ttk.Label(
            master=help_frame_5,
            text="""

            After all fields have been filled out, clicking the 'Publish Info' button at the bottom right of the GUI will publish the
            now determined data to the indicated destination archive file. 

            If the 'Log File Destination' section above was filled out before 'Publish Info' was pressed, the source log file will be
            moved to the designated directory after data publishing is complete.

            """,
        )
        publishing_data_label.grid(row=0, column=0, padx=5, pady=5)

        other_label = ttk.Label(
            master=help_frame_6,
            text="""

            Google Sheets functionality is only designated for  use by the lab operator, and should not be used otherwise. Proper access
            authentication is required for use and will not be operational otherwise.

            For input fields, most input should be self explanatory with specifics laid out below:
            - Proposal # -> should be established beforehand, known to the users before scanning
            - Name -> The names of each person associated with the project the scan is linked with
            - Date -> The date of the scan
            - Position -> The position(s) of the users from the aforementioned Names list
            - Operator -> The specific person who operated the machine for the designated scan
            - Internal/External/Both -> If the scan is for an NYIT affiliated group, External affiliated group, or both
            - Filename -> The name of the log file, not including extension
            - Voltage -> The voltage used during the scan
            - Current -> The current used during the scan
            - Resolution -> The resolution used during the scan
            - Filter -> The filter used during the scan
            - Exposure -> The exposure specified during the scan
            - Pixel Size -> The pixel size designated during the scan
            - Image Format -> The image format output by the scan
            - Type of Scan -> The type of scan performed (i.e. Regular, Oversized, Segment #, etc.)
            - Rotation Step -> The rotation step designated during the scan
            - Frames -> The amount of frames designated during the scan
            - Random Movement -> If there was random movement designated during the scan
            - Rotate 360 -> If there was 360 degree movement during the scan
            - Scan Duration -> The duration of the indicated scan
            - Comments -> Any comments about the scan that need to be recorded for further clarification
            - Log File Archived -> If the log file was archived (or will be archived) after the data has been published
            - Added to Viscenter Storage -> If the data was added to the designated Visualization Center Storage hub

            """,
        )
        other_label.grid(row=0, column=0, padx=5, pady=5)

    def select_destination_file(self):
        """opens a file dialog asking the user to select an
        archive file where the data will be exported to"""
        selected_destination_file = fdlg.askopenfilename()
        self.select_destination_file_label.configure(
            text=str(selected_destination_file)
        )
        if str(selected_destination_file) != "":
            self.indicator_light_destination_file.itemconfig(
                self.import_data_light, fill="green"
            )
            self.target_archive_file = selected_destination_file
        # print(selected_destination_file)

    def select_log_source_file(self):
        """opens a file dialog asking the user to select a log file
        from which the data to be exported will be extracted from"""
        self.selected_log_source_file = fdlg.askopenfilename()
        self.select_log_file_label.configure(text=str(self.selected_log_source_file))
        if str(self.selected_log_source_file) != "":
            self.indicator_light_importdata.itemconfig(
                self.import_data_light, fill="yellow"
            )
        # print(selected_log_source_file)

    def parse_data(self):
        """parses through the data extracted from the log file, filling all
        possible fields on the GUI for checking and editing"""
        log_source_file_open = open(
            str(self.selected_log_source_file), "r", encoding="utf-8"
        )
        log_source_file_text = log_source_file_open.read()
        log_source_file_open.close()
        self.log_file_text = log_source_file_text
        month_dict = {
            "Jan": 1,
            "Feb": 2,
            "Mar": 3,
            "Apr": 4,
            "May": 5,
            "Jun": 6,
            "Jul": 7,
            "Aug": 8,
            "Sep": 9,
            "Oct": 10,
            "Nov": 11,
            "Dec": 12,
        }
        log_date = re.compile(
            r"Study Date and Time\s*=\s*(\w+)\s*(\d\d),\s*(\d\d\d\d).*", re.I
        )
        log_filename = re.compile(r"Filename Prefix\s*=\s*(.*)", re.I)
        log_voltage = re.compile(r"Source Voltage \(kV\)\s*=\s*(\d+)", re.I)
        log_current = re.compile(r"Source Current \(uA\)\s*=\s*(\d+)", re.I)
        log_resolution = re.compile(r"Number of Rows\s*=\s*(\d+)", re.I)
        log_exposure = re.compile(r"Exposure \(ms\)\s*=\s*(\d+)", re.I)
        log_pixelsize = re.compile(r"Image Pixel Size\s*\(um\)\s*=(\d+\.\d+)", re.I)
        log_imageformat = re.compile(r"Image Format\s*=\s*(\w+)", re.I)
        log_rotationstep = re.compile(r"Rotation Step\s*\(deg\)\s*=\s*(\d+\.\d+)", re.I)
        log_frames = re.compile(r"Frame Averaging\s*=[ON|OFF]{2,3}\s*\((\d+)\)", re.I)
        log_randommovement = re.compile(
            r"Random Movement\s*=[ON|OFF]{2,3}\s*\((\d+)\)", re.I
        )
        log_360 = re.compile(r"Use 360 Rotation\s*=\s*(\w+)", re.I)
        log_scanduration = re.compile(r"Scan duration\s*=\s*(\d{2}:\d{2}:\d{2})", re.I)

        try:
            date_find = log_date.search(str(self.log_file_text))
            if date_find is not None:
                # print("Date found")
                self.date_textvar.set(
                    str(month_dict[str(date_find.group(1))])
                    + "/"
                    + str(date_find.group(2))
                    + "/"
                    + str(date_find.group(3))
                )
                # print(str(date_find.group(2)))
            else:
                # print("Date not found???")
                pass
        except Exception:
            date_find = None
            print("Date not found. Fix and try again.")
        try:
            filename_find = log_filename.search(str(self.log_file_text))
            if filename_find is not None:
                self.filename_textvar.set(str(filename_find.group(1)))
        except Exception:
            filename_find = None
            print("File Name not found. Fix and try again.")
        try:
            voltage_find = log_voltage.search(str(self.log_file_text))
            if voltage_find is not None:
                self.voltage_textvar.set(str(voltage_find.group(1)))
        except Exception:
            voltage_find = None
            print("Voltage not found. Fix and try again.")
        try:
            current_find = log_current.search(str(self.log_file_text))
            if current_find is not None:
                self.current_textvar.set(str(current_find.group(1)))
        except Exception:
            current_find = None
            print("Current not found. Fix and try again.")
        try:
            resolution_find = log_resolution.search(str(self.log_file_text))
            if resolution_find is not None:
                temp_res = str(resolution_find.group(1))
                act_res = None
                if int(temp_res) == 2240:
                    act_res = "2K"
                elif int(temp_res) == 1120:
                    act_res = "1K"
                elif int(temp_res) == 560:
                    act_res = "0.5K"
                else:
                    act_res = "None"
                self.resolution_textvar.set(act_res)
        except Exception:
            resolution_find = None
            print("Resolution not found. Fix and try again.")
        try:
            exposure_find = log_exposure.search(str(self.log_file_text))
            if exposure_find is not None:
                self.exposure_textvar.set(str(exposure_find.group(1)))
        except Exception:
            exposure_find = None
            print("Exposure not found. Fix and try again.")
        try:
            pixelsize_find = log_pixelsize.search(str(self.log_file_text))
            if pixelsize_find is not None:
                self.pixelsize_textvar.set(str(pixelsize_find.group(1)))
        except Exception:
            pixelsize_find = None
            print("Pixel Size not found. Fix and try again.")
        try:
            imageformat_find = log_imageformat.search(str(self.log_file_text))
            if imageformat_find is not None:
                self.imageformat_textvar.set(str(imageformat_find.group(1)))
        except Exception:
            imageformat_find = None
            print("Image Format not found. Fix and try again.")
        try:
            rotationstep_find = log_rotationstep.search(str(self.log_file_text))
            if rotationstep_find is not None:
                self.rotationstep_textvar.set(str(rotationstep_find.group(1)))
        except Exception:
            rotationstep_find = None
            print("Rotation Step not found. Fix and try again.")
        try:
            frames_find = log_frames.search(str(self.log_file_text))
            if frames_find is not None:
                self.frames_textvar.set(str(frames_find.group(1)))
        except Exception:
            frames_find = None
            print("Frames not found. Fix and try again.")
        try:
            randommovement_find = log_randommovement.search(str(self.log_file_text))
            if randommovement_find is not None:
                self.randommovement_textvar.set(str(randommovement_find.group(1)))
        except Exception:
            randommovement_find = None
            print("Random Movement not found. Fix and try again.")
        try:
            move360_find = log_360.search(str(self.log_file_text))
            if str(move360_find).capitalize() == "YES":
                self.rotate360_yesno.set(True)
            else:
                self.rotate360_yesno.set(False)
        except Exception:
            move360_find = None
            print("360 Movement not found. Fix and try again.")
        try:
            scanduration_find = log_scanduration.search(str(self.log_file_text))
            if scanduration_find is not None:
                self.scanduration_textvar.set(str(scanduration_find.group(1)))
        except Exception:
            scanduration_find = None
            print("Scan Duration not found. Fix and try again.")

        self.indicator_light_importdata.itemconfig(self.import_data_light, fill="green")

    def shift_log_file(self):
        """shift the log file from its current location to a desired end directory"""
        self.target_dir = str(fdlg.askdirectory())
        self.shift_log_file_label.configure(text=str(self.target_dir))
        if str(self.target_dir) != "":
            self.indicator_light_shift_logfile.itemconfig(
                self.shift_log_file_light, fill="green"
            )

    def publish_info(self):
        """take the entered and parsed data and export it to the desired archive file"""
        self.data_dict["Proposal Number"] = self.proposalnumber_textvar.get()
        self.data_dict["Name"] = self.name_textvar.get()
        self.data_dict["Date"] = self.date_textvar.get()
        self.data_dict["Position"] = self.position_textvar.get()
        self.data_dict["Operator"] = self.operator_textvar.get()
        self.data_dict["NYIT or External"] = self.nyitorext_textvar.get()
        self.data_dict["Filename"] = self.filename_textvar.get()
        self.data_dict["Voltage"] = self.voltage_textvar.get()
        self.data_dict["Current"] = self.current_textvar.get()
        temp_res = self.resolution_textvar.get()
        act_res = None
        if (temp_res != "0.5K") and (temp_res != "1K") and (temp_res != "2K"):
            if int(temp_res) == 2240:
                act_res = "2K"
            elif int(temp_res) == 1120:
                act_res = "1K"
            elif int(temp_res) == 560:
                act_res = "0.5K"
            else:
                act_res = "None"
        else:
            act_res = str(temp_res)
        self.data_dict["Resolution"] = act_res
        self.data_dict["Filter"] = self.filter_textvar.get()
        self.data_dict["Exposure"] = self.exposure_textvar.get()
        self.data_dict["Pixel Size"] = self.pixelsize_textvar.get()
        self.data_dict["Image Format"] = self.imageformat_textvar.get()
        self.data_dict["Type of Scan"] = self.typeofscan_textvar.get()
        self.data_dict["Rotation Step"] = self.rotationstep_textvar.get()
        self.data_dict["Frames"] = self.frames_textvar.get()
        self.data_dict["Random Movement"] = self.randommovement_textvar.get()
        if self.rotate360_yesno.get() == 1:
            self.data_dict["Rotate 360"] = "yes"
        else:
            self.data_dict["Rotate 360"] = "no"
        self.data_dict["Scan Duration"] = self.scanduration_textvar.get()
        self.data_dict["Comments"] = self.comments_input.get(
            "1.0", "end-1c"
        )  # Deletes the newline character that's added with getting the input
        if self.logfilearchive_yesno.get() == 1:
            self.data_dict["Log File Archived"] = "yes"
        else:
            self.data_dict["Log File Archived"] = "no"
        if self.visstorage_yesno.get() == 1:
            self.data_dict["Viscenter Storage"] = "yes"
        else:
            self.data_dict["Viscenter Storage"] = "no"
        appendable_data = [
            self.data_dict["Proposal Number"],
            self.data_dict["Name"],
            self.data_dict["Date"],
            self.data_dict["Position"],
            self.data_dict["NYIT or External"],
            self.data_dict["Filename"],
            self.data_dict["Voltage"],
            self.data_dict["Current"],
            self.data_dict["Resolution"],
            self.data_dict["Filter"],
            self.data_dict["Exposure"],
            self.data_dict["Pixel Size"],
            self.data_dict["Image Format"],
            self.data_dict["Type of Scan"],
            self.data_dict["Rotation Step"],
            self.data_dict["Frames"],
            self.data_dict["Random Movement"],
            self.data_dict["Rotate 360"],
            self.data_dict["Scan Duration"],
            self.data_dict["Comments"],
            self.data_dict["Log File Archived"],
            self.data_dict["Viscenter Storage"],
            self.data_dict["Operator"],
        ]
        # df = pd.DataFrame(
        #     np.array([appendable_data]),
        # )

        if self.selectedfile_radbuttonvar.get() == 1:  # Excel
            # Check if target excel sheet is selected
            myxlfile = Path(self.target_archive_file)
            wb = openpyxl.load_workbook(myxlfile)
            sheet = wb[wb.sheetnames[0]]
            sheetmaxrow = sheet.max_row + 1
            for i in range(sheet.max_column):
                column_let = get_column_letter(i + 1)
                cell2edit = str(column_let) + str(sheetmaxrow)
                if 0 <= i < len(appendable_data):
                    sheet[cell2edit] = appendable_data[i]
            wb.save(myxlfile)

            # Move file after publishing data
            if str(self.selected_log_source_file) != "":
                shutil.copy(str(self.selected_log_source_file), str(self.target_dir))
        else:  # Google Sheets
            pass

    def reset_clear(self):
        """clear all user input and reset all fields for re-filling"""
        if self.destinationfilekeep_yesno.get() != 1:
            self.select_destination_file_label.configure(text=str(""))
            self.indicator_light_destination_file.itemconfig(
                self.import_data_light, fill="red"
            )
        else:
            pass

        self.select_log_file_label.configure(text=str(""))
        self.indicator_light_importdata.itemconfig(self.import_data_light, fill="red")

        if self.destinationdirectorykeep_yesno.get() != 1:
            self.shift_log_file_label.configure(text=str(""))
            self.indicator_light_shift_logfile.itemconfig(
                self.shift_log_file_light, fill="red"
            )
        else:
            pass

        self.proposalnumber_textvar.set("Enter Here")
        self.name_textvar.set("Enter Here")
        self.date_textvar.set("Known")
        self.position_textvar.set("Enter Here")
        self.operator_textvar.set("Enter Here")
        self.nyitorext_textvar.set("Enter Here")
        self.filename_textvar.set("Known")
        self.voltage_textvar.set("Known")
        self.current_textvar.set("Known")
        self.resolution_textvar.set("Click to Select")
        self.filter_textvar.set("Enter or Select Here")
        self.exposure_textvar.set("Known")
        self.pixelsize_textvar.set("Known")
        self.imageformat_textvar.set("Known")
        self.typeofscan_textvar.set("Enter Here")
        self.rotationstep_textvar.set("Known")
        self.frames_textvar.set("Known")
        self.randommovement_textvar.set("Known")
        self.rotate360_yesno.set(0)
        self.scanduration_textvar.set("Known")
        self.comments_input.delete(1.0, tk.END)
        self.comments_input.insert(tk.END, "Enter Here")
        self.logfilearchive_yesno.set(0)
        self.visstorage_yesno.set(0)


if __name__ == "__main__":
    root = MainWindow()  # Create a MainWindow Object labeled root
    root.mainloop()  # Run root through the mainloop
